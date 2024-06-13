import mysql.connector
import openpyxl
mydb = mysql.connector.connect(
    host="localhost",
    user="root",
    password="Haris@51",
    database="mydatabase"
)
mycursor = mydb.cursor()
mycursor.execute("SELECT * FROM tbl_form_detail")
myresult = mycursor.fetchall()
column_names = [i[0] for i in mycursor.description]
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Form Details"
for col_num, column_title in enumerate(column_names, 1):
    ws.cell(row=1, column=col_num, value=column_title)
for row_num, row_data in enumerate(myresult, 2):
    for col_num, cell_value in enumerate(row_data, 1):
        ws.cell(row=row_num, column=col_num, value=cell_value)
wb.save("FormDetails.xlsx")
print("Data has been written to FormDetails.xlsx")
